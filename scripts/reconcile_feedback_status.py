"""
Task #96 / T002 — Reconcile feedback workbook status with merged work.

Reads attached_assets/KONTi_Dashboard_Feedback_Consolidated_v2_1777518178155.xlsx
Writes attached_assets/KONTi_Dashboard_Feedback_Consolidated_v3_addressed.xlsx
Writes reports/feedback-status-reconciled.md

Mapping is hand-curated from listProjectTasks() (see commit message + plan file).
"""
import openpyxl
from copy import copy
from pathlib import Path

SRC = "attached_assets/KONTi_Dashboard_Feedback_Consolidated_v2_1777518178155.xlsx"
DST = "attached_assets/KONTi_Dashboard_Feedback_Consolidated_v3_addressed.xlsx"
REPORT = "reports/feedback-status-reconciled.md"

# (status, note appended to col 14 "Scope Rationale")
# Status values: Done | In Progress | Open | Needs Decision | Needs Spec
MAP = {
    # --- A. Project Detail ---
    "A-01": ("Done", "Done in #99 (Reviewer feedback bundle #2): project-invoices.tsx Total/Paid/Balance/Status columns now render from invoice data."),
    "A-02": ("Open", None),
    "A-03": ("Done", "Done in #61 (client portal expansion: client uploads enabled)."),
    "A-04": ("Done", "Done in #63 (document organization: contracts/agreements grouping)."),
    "A-05": ("Open", None),
    "A-06": ("Done", "Done in #61 (per-document client visibility) and reinforced by #88 (client ownership checks)."),
    "A-07": ("Open", "Tracked as Task #47 (Site photo upload, categorization, links from report)."),
    "A-08": ("Done", "Done in #75 (ClientContactCard with phone, postal, physical addresses)."),
    "A-09": ("Open", None),
    "A-10": ("Needs Decision", None),
    "A-11": ("Done — needs verification", "Likely closed by #62 + #75 (Contractor Estimate Rollup on the project report); needs PM eyes-on confirmation that the consolidated view matches the original ask."),
    "A-12": ("Open", "Admin-side audit log shipped in #73; client-side audit log still V2."),
    "A-13": ("Done", "Done in #62 (KONTi brand pass) and #74 (header text readable on bright cover photos)."),

    # --- B. Cost Calculator ---
    "B-01": ("Done", "Done in #75 (CSV header aliases: Description, UnitPrice, etc.)."),
    "B-02": ("Open", None),
    "B-03": ("Done", "Done in #75 (calculator auto-populates from imported materials)."),
    "B-04": ("Done", "Done in #75 (inline edit + PATCH /projects/:id/calculations/:lineId persistence)."),
    "B-05": ("Done", "Done in #75 (Project Information panel with bathrooms/kitchens/margin/mgmt-fee inputs)."),
    "B-06": ("Done", "Done in #99 (Reviewer feedback bundle #2): calculator estimate table now groups by category with per-category subtotal cards, mirroring the team's external estimate format."),
    "B-07": ("Done", "Done in #99 (Reviewer feedback bundle #2): renamed Imports tab to 'Imported Materials' / 'Materiales Importados' with hover tooltip describing CSV/Excel bulk import."),
    "B-08": ("Done", "Done in #75 (renamed to 'Effective Price' with tooltip + legend)."),
    "B-09": ("Done", "Done in #99 (Reviewer feedback bundle #2): added 'Receipts & Variance' shortcut card on dashboard linking team users straight to /calculator?tab=variance."),
    "B-10": ("Done", "Done in #27 (receipts and contractor estimates persist across restart)."),
    "B-11": ("Done", "Done in #28 (real PDF/image OCR replaces the mock)."),
    "B-12": ("Done", "Done in #29 (PDF export now uses the saved report template)."),
    "B-13": ("Done", "Done in #99 (Reviewer feedback bundle #2): Materials Library 'Add Material' button now opens a modal that POSTs a single material via the existing /api/estimating/materials/import endpoint and refreshes the catalog."),
    "B-14": ("Open", None),

    # --- C. Project Report ---
    "C-01": ("Open", "Punchlist persistence shipped in #32; photo links + categories on the report still pending."),
    "C-02": ("Done", "Done in #99 (Reviewer feedback bundle #2): contractor BOM detail now gated by !isClientView so client viewers only see the Cost-by-Category rollup and never the raw line items."),
    "C-03": ("Done", "Done in #99 (Reviewer feedback bundle #2): phase numbers no longer rendered anywhere in the project report (phase chips, timeline, donut all show labels only)."),
    "C-04": ("Done", "Done in #99 (Reviewer feedback bundle #2): added Phase Progress donut on the project report mirroring the punchlist phase-pie style with per-phase % completion and an avg-completion centre label."),
    "C-05": ("Done", "Done in #99 (Reviewer feedback bundle #2): renamed 'Site Conditions' to 'Weather Status' / 'Estado del Clima' in the report header tile and the dedicated weather section."),
    "C-06": ("Done", "Done in #99 (Reviewer feedback bundle #2): Cost-by-Category card and the BOM detail are both driven from the same calc.subtotalByCategory data so totals always match."),
    "C-07": ("Done", "Done in #75 (mgmt fee editable from the project report; flows through to the rollup)."),
    "C-08": ("Done", "Done in #71 (P1 quick wins: report logo enlarged)."),
    "C-09": ("Done", "Done in #62 (KONTi brand pass replaced the dark/black palette)."),
    "C-10": ("Done", "Done in #99 (Reviewer feedback bundle #2): replaced auto-generated reportDate with an editable <input type='date'> in the sticky report header, persisted per project via localStorage."),
    "C-11": ("Open", "Tracked as Task #47 (Site photo upload, categorization, links from report)."),
    "C-12": ("Done", "Done in #62 (light backgrounds across the project report)."),

    # --- D. AI Assistant ---
    "D-01": ("Done", "Done in #30 (AI assistant notes/updates persist across restart)."),
    "D-02": ("Needs Decision", None),

    # --- E. Permits ---
    "E-01": ("Open", "Tracked as Task #48 (Permits page: legal header + split by permit type)."),
    "E-02": ("Open", "Tracked as Task #48 (Permits page: legal header + split by permit type)."),
    "E-03": ("Done", "Done in #71 (P1 quick wins: permits copy fixed)."),
    "E-04": ("Needs Decision", None),
    "E-05": ("Needs Decision", None),

    # --- F. Dashboard ---
    "F-01": ("Done", "Done in #71 (P1 quick wins: clickable activity)."),
    "F-02": ("Done", "Done in #61 (client home in client portal) and #72 (dashboard restructure)."),

    # --- G. Team Directory ---
    "G-01": ("Open", None),

    # --- H. Leads / CRM ---
    "H-01": ("Done", "Done in #99 (Reviewer feedback bundle #2): leads page now renders an inline lead-score legend (Hot / Warm / Cold / New thresholds) right next to the table."),
    "H-02": ("Needs Decision", None),
    "H-03": ("Needs Decision", None),

    # --- I. Demo Project ---
    "I-01": ("Done", "Done in #60 (file upload regression on the demo project fixed)."),
    "I-02": ("Done", "Done in #99 (Reviewer feedback bundle #2): document upload modal now requires a category dropdown so demo-project docs are sorted into the correct buckets."),
    "I-03": ("Done", "Done in #32 (punchlist persists across restart)."),
    "I-04": ("Open", None),

    # --- J. Drive ---
    "J-01": ("Needs Decision", None),
}

ALL_IDS = set(MAP.keys())


def update_row_status(ws, row_idx, new_status, note):
    """Update Status (col 12) and append note to Scope Rationale (col 14)."""
    status_cell = ws.cell(row_idx, 12)
    status_cell.value = new_status
    if note:
        rationale_cell = ws.cell(row_idx, 14)
        existing = (rationale_cell.value or "").strip()
        if note not in existing:
            rationale_cell.value = f"{existing} | {note}".lstrip(" |") if existing else note


def reconcile_sheet(ws, observed):
    """Walk a sheet and update statuses for any matching ID rows."""
    for r in range(1, ws.max_row + 1):
        id_ = ws.cell(r, 1).value
        if not id_:
            continue
        id_ = str(id_).strip()
        if id_ in ALL_IDS:
            new_status, note = MAP[id_]
            old_status = ws.cell(r, 12).value
            update_row_status(ws, r, new_status, note)
            observed.append((id_, old_status, new_status))


def refresh_summary(ws, sheet1):
    """Recompute Status counts on the Summary sheet."""
    counts = {"Open": 0, "In Progress": 0, "Done": 0, "Done — needs verification": 0,
              "Needs Spec": 0, "Needs Decision": 0}
    for r in range(1, sheet1.max_row + 1):
        id_ = sheet1.cell(r, 1).value
        if not id_:
            continue
        id_ = str(id_).strip()
        if id_ in ALL_IDS:
            s = sheet1.cell(r, 12).value
            if s in counts:
                counts[s] += 1

    label_to_row = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label in counts:
            label_to_row[label] = r
    for label, count in counts.items():
        if label in label_to_row:
            # Use explicit numeric 0 (not blank) so the totals are unambiguous.
            ws.cell(label_to_row[label], 2).value = count

    # Add a "Done — needs verification" row to the summary if it isn't already
    # listed (the original v2 sheet only had Done / Needs Spec / Needs Decision).
    if "Done — needs verification" not in label_to_row:
        # Append immediately after the last status row (row 21 in the v2 layout).
        # Find the last row in the "By Status" block (between r16 header and the
        # next blank/section break) and insert there.
        target = max(label_to_row.values()) + 1
        ws.cell(target, 1).value = "Done — needs verification"
        ws.cell(target, 2).value = counts["Done — needs verification"]

    return counts


def main():
    wb = openpyxl.load_workbook(SRC)
    sheet1 = wb["KONTi Dashboard Feedback"]
    sheet2 = wb["V2 Backlog"]
    summary = wb["Summary"]

    observed = []
    reconcile_sheet(sheet1, observed)
    reconcile_sheet(sheet2, observed)

    counts = refresh_summary(summary, sheet1)

    Path(DST).parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    print(f"Wrote {DST}")
    print(f"Status totals: {counts}")
    print(f"Total mapped IDs: {len(ALL_IDS)} (sheet1 rows seen: {len(observed)})")

    write_report(observed, counts)


def write_report(observed, counts):
    by_status = {}
    for id_, old, new in observed:
        by_status.setdefault(new, []).append((id_, old))
    for k in by_status:
        by_status[k].sort()

    lines = []
    lines.append("# Feedback status reconciliation — Apr 30 2026")
    lines.append("")
    lines.append("Source workbook: `attached_assets/KONTi_Dashboard_Feedback_Consolidated_v2_1777518178155.xlsx`")
    lines.append("")
    lines.append("Reconciled workbook: `attached_assets/KONTi_Dashboard_Feedback_Consolidated_v3_addressed.xlsx`")
    lines.append("")
    lines.append("## Totals (Sheet 1, all 57 V1+V2 items)")
    lines.append("")
    lines.append("| Status | Count |")
    lines.append("|---|---:|")
    for k in ["Open", "In Progress", "Done", "Done — needs verification", "Needs Spec", "Needs Decision"]:
        lines.append(f"| {k} | {counts.get(k,0)} |")
    lines.append("")
    lines.append("## Items moved to **Done — needs verification**")
    lines.append("")
    lines.append("These rows look closed on paper but a PM should eyeball the live UI before promoting them to plain Done.")
    lines.append("")
    lines.append("| ID | Was | Now | Why verification is suggested |")
    lines.append("|---|---|---|---|")
    seen = set()
    for id_, old, new in sorted(observed):
        if new == "Done — needs verification" and id_ not in seen:
            seen.add(id_)
            note = MAP[id_][1] or ""
            lines.append(f"| {id_} | {old or '—'} | {new} | {note} |")
    lines.append("")
    lines.append("## Items moved to **Done**")
    lines.append("")
    lines.append("| ID | Was | Now | Closed by |")
    lines.append("|---|---|---|---|")
    seen = set()
    for id_, old, new in sorted(observed):
        if new == "Done" and id_ not in seen:
            seen.add(id_)
            note = MAP[id_][1] or ""
            lines.append(f"| {id_} | {old or '—'} | Done | {note} |")
    lines.append("")
    lines.append("## Items still **Open**")
    lines.append("")
    lines.append("| ID | Was | Now | Note |")
    lines.append("|---|---|---|---|")
    seen = set()
    for id_, old, new in sorted(observed):
        if new == "Open" and id_ not in seen:
            seen.add(id_)
            note = MAP[id_][1] or "—"
            lines.append(f"| {id_} | {old or '—'} | Open | {note} |")
    lines.append("")
    lines.append("## Items needing a product decision")
    lines.append("")
    lines.append("| ID | Was | Now |")
    lines.append("|---|---|---|")
    seen = set()
    for id_, old, new in sorted(observed):
        if new == "Needs Decision" and id_ not in seen:
            seen.add(id_)
            lines.append(f"| {id_} | {old or '—'} | Needs Decision |")
    lines.append("")
    lines.append("## Notes")
    lines.append("")
    lines.append("- Sheet 4 (Legend & Guide) is preserved unchanged.")
    lines.append("- Sheet 2 (V2 Backlog) statuses are kept in sync with Sheet 1 for the same IDs (B-11, D-01, etc.).")
    lines.append("- 'Done' rows have a one-line justification appended to the Scope Rationale column linking to the merged task ref.")
    lines.append("- Items A-07 / C-11 / E-01 / E-02 are proposed but not yet merged tasks (#47, #48); they remain Open.")

    Path(REPORT).parent.mkdir(parents=True, exist_ok=True)
    Path(REPORT).write_text("\n".join(lines) + "\n")
    print(f"Wrote {REPORT}")


if __name__ == "__main__":
    main()
