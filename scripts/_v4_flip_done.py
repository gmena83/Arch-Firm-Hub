"""
Task #119 — Mirror the A-07 / E-01 / E-02 status flips into the v4 workbook
that the dashboard's status report references, with bilingual EN | ES
verification notes matching the v4 format. C-11 and G-01 are already 'Done'
in v4, so we leave them alone.

This is a one-shot helper. The reconcile_feedback_status.py script remains
the authoritative source of truth for the v3 workbook + markdown report;
this file just keeps v4 in sync.
"""
import openpyxl
from pathlib import Path

V4 = "attached_assets/reports/KONTi_Dashboard_Feedback_Consolidated_v4.xlsx"

FLIPS = {
    "A-07": (
        "Done",
        "EN: Done in #105 — Photos & Media tab now lives on the project detail "
        "with bulk site-photo upload, category tags, and a per-project gallery "
        "rendered into the project report. File: "
        "artifacts/konti-dashboard/src/components/site-photos-gallery.tsx + "
        "artifacts/konti-dashboard/src/pages/project-detail.tsx. "
        "| ES: Hecho en #105 — la pestaña Fotos & Medios ya existe en el detalle "
        "del proyecto con carga masiva de fotos de obra, etiquetas por categoría "
        "y una galería por proyecto incluida en el reporte. Archivo: "
        "artifacts/konti-dashboard/src/components/site-photos-gallery.tsx + "
        "artifacts/konti-dashboard/src/pages/project-detail.tsx."
    ),
    "E-01": (
        "Done",
        "EN: Done in #106 — Permits page now renders the legal/engineer header "
        "block above the list, mirroring the team's spreadsheet. File: "
        "artifacts/konti-dashboard/src/pages/permits.tsx. "
        "| ES: Hecho en #106 — la página de Permisos ahora muestra el bloque "
        "de encabezado legal/ingeniero arriba del listado, replicando la hoja "
        "de cálculo del equipo. Archivo: "
        "artifacts/konti-dashboard/src/pages/permits.tsx."
    ),
    "E-02": (
        "Done",
        "EN: Done in #106 — Permits are now grouped by type (PCOC, USO, "
        "Consulta de Ubicación, etc.) with separate sections per family, "
        "matching the team's permit Excel. File: "
        "artifacts/konti-dashboard/src/pages/permits.tsx. "
        "| ES: Hecho en #106 — los Permisos ahora están agrupados por tipo "
        "(PCOC, USO, Consulta de Ubicación, etc.) con secciones separadas por "
        "familia, replicando el Excel de permisos del equipo. Archivo: "
        "artifacts/konti-dashboard/src/pages/permits.tsx."
    ),
}

ID_COL = 1
STATUS_COL = 12
VERIF_COL = 15


def flip_sheet(ws):
    flipped = []
    for r in range(2, ws.max_row + 1):
        id_ = ws.cell(r, ID_COL).value
        if not id_:
            continue
        id_ = str(id_).strip()
        if id_ in FLIPS:
            new_status, new_verif = FLIPS[id_]
            old_status = ws.cell(r, STATUS_COL).value
            ws.cell(r, STATUS_COL).value = new_status
            ws.cell(r, VERIF_COL).value = new_verif
            flipped.append((id_, old_status, new_status))
    return flipped


def refresh_summary(ws_summary, ws_main):
    counts = {}
    for r in range(2, ws_main.max_row + 1):
        s = ws_main.cell(r, STATUS_COL).value
        if s:
            counts[s] = counts.get(s, 0) + 1

    label_to_row = {}
    for r in range(1, ws_summary.max_row + 1):
        label = ws_summary.cell(r, 1).value
        if isinstance(label, str) and label in counts:
            label_to_row[label] = r

    for label, count in counts.items():
        if label in label_to_row:
            ws_summary.cell(label_to_row[label], 2).value = count
    return counts


def main():
    path = Path(V4)
    wb = openpyxl.load_workbook(path)
    main_ws = wb["KONTi Dashboard Feedback"]
    flipped_main = flip_sheet(main_ws)
    print(f"Main sheet flips: {flipped_main}")

    if "V2 Backlog" in wb.sheetnames:
        flipped_v2 = flip_sheet(wb["V2 Backlog"])
        print(f"V2 Backlog flips: {flipped_v2}")

    if "Summary" in wb.sheetnames:
        counts = refresh_summary(wb["Summary"], main_ws)
        print(f"Summary counts: {counts}")

    wb.save(path)
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
